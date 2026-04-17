# -*- coding: utf-8 -*-
"""
Load in time data from Rattlesnake runs
"""
# Copyright 2022 National Technology & Engineering Solutions of Sandia,
# LLC (NTESS). Under the terms of Contract DE-NA0003525 with NTESS, the U.S.
# Government retains certain rights in this software.

# This program is free software: you can redistribute it and/or modify
# it under the terms of the GNU General Public License as published by
# the Free Software Foundation, either version 3 of the License, or
# (at your option) any later version.

# This program is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
# GNU General Public License for more details.

# You should have received a copy of the GNU General Public License
# along with this program.  If not, see <https://www.gnu.org/licenses/>.

from __future__ import annotations
import netCDF4 as nc4
import numpy as np
from ..core.sdynpy_coordinate import coordinate_array, outer_product, CoordinateArray, _string_map
from ..core.sdynpy_data import (
    data_array,
    FunctionTypes,
    time_history_array,
    TimeHistoryArray,
    power_spectral_density_array,
    PowerSpectralDensityArray,
    transfer_function_array,
    TransferFunctionArray,
    coherence_array,
    CoherenceArray,
    multiple_coherence_array,
    MultipleCoherenceArray,
)
from ..core.sdynpy_system import System
from ..core.sdynpy_matrix import matrix
import pandas as pd
import sys
import openpyxl as opxl
import os
import warnings


def read_rattlesnake_output(
    file,
    coordinate_override_column=None,
    read_only_indices=None,
    read_variable="time_data",
    abscissa_start=None,
    abscissa_stop=None,
    downsample=None,
):
    """
    Reads in an nc4 Rattlesnake data file and returns the time history array as well
    as the channel table

    Parameters
    ----------
    file : str or netCDF4.Dataset
        Path to the nc4 file to read, or an already-open ``netCDF4.Dataset``.
    coordinate_override_column : str, optional
        Name of a channel-table column whose string values are parsed as
        coordinates.  When not specified, coordinates are assembled from the
        ``node_number`` and ``node_direction`` channel-table columns.
    read_only_indices : slice or iterable, optional
        A valid indexing operation to select which channel indices to read
    read_variable : str, optional
        The time variable from the Rattlesnake file to read.  These will
        generally be time_data, time_data_1, time_data_2, etc. depending on
        how many streams exist in the file.  The default is 'time_data'.
    abscissa_start : float, optional
        Data will not be extracted for abscissa values less than this value
    abscissa_stop : float, optional
        Data will not be extracted for abscissa values greater than this value
    downsample : int, optional
        A step size to use to downsample the dataset when reading

    Returns
    -------
    data_array : TimeHistoryArray
        Time history data in the Rattlesnake output file
    channel_table : DataFrame
        Pandas Dataframe containing the channel table information

    """
    if isinstance(file, str):
        ds = nc4.Dataset(file, "r")
    elif isinstance(file, nc4.Dataset):
        ds = file
    if read_only_indices is None:
        read_only_indices = slice(None)
    if abscissa_start is None:
        start_index = None
    else:
        start_index = int(np.ceil(abscissa_start * ds.sample_rate))
    if abscissa_stop is None:
        stop_index = None
    else:
        stop_index = int(np.ceil(abscissa_stop * ds.sample_rate))
    abscissa_slice = slice(start_index, stop_index, downsample)
    output_data = np.array(ds[read_variable][:, abscissa_slice][read_only_indices])
    abscissa = (
        np.arange(
            0 if start_index is None else start_index,
            ds[read_variable].shape[-1] if stop_index is None else stop_index,
            1 if downsample is None else downsample,
        )
        / ds.sample_rate
    )
    if coordinate_override_column is None:
        nodes = [
            int("".join(char for char in node if char in "0123456789"))
            for node in ds["channels"]["node_number"][...][read_only_indices]
        ]
        directions = np.array(ds["channels"]["node_direction"][...][read_only_indices], dtype="<U3")
        coordinates = coordinate_array(nodes, directions)[:, np.newaxis]
    else:
        coordinates = coordinate_array(
            string_array=ds["channels"][coordinate_override_column][read_only_indices]
        )[:, np.newaxis]
    array = {name: np.array(variable[:]) for name, variable in ds["channels"].variables.items()}
    channel_table = pd.DataFrame(array)
    comment1 = np.char.add(
        np.char.add(
            np.array(ds["channels"]["channel_type"][...][read_only_indices], dtype="<U80"),
            np.array(" :: "),
        ),
        np.array(ds["channels"]["unit"][...][read_only_indices], dtype="<U80"),
    )
    comment2 = np.char.add(
        np.char.add(
            np.array(ds["channels"]["physical_device"][...][read_only_indices], dtype="<U80"),
            np.array(" :: "),
        ),
        np.array(ds["channels"]["physical_channel"][...][read_only_indices], dtype="<U80"),
    )
    comment3 = np.char.add(
        np.char.add(
            np.array(ds["channels"]["feedback_device"][...][read_only_indices], dtype="<U80"),
            np.array(" :: "),
        ),
        np.array(ds["channels"]["feedback_channel"][...][read_only_indices], dtype="<U80"),
    )
    comment4 = np.array(ds["channels"]["comment"][...][read_only_indices], dtype="<U80")
    comment5 = np.array(ds["channels"]["make"][...][read_only_indices], dtype="<U80")
    for key in ("model", "serial_number", "triax_dof"):
        comment5 = np.char.add(comment5, np.array(" "))
        comment5 = np.char.add(
            comment5, np.array(ds["channels"][key][...][read_only_indices], dtype="<U80")
        )
    time_data = data_array(
        FunctionTypes.TIME_RESPONSE,
        abscissa,
        output_data,
        coordinates,
        comment1,
        comment2,
        comment3,
        comment4,
        comment5,
    )
    if isinstance(file, str):
        ds.close()
    return time_data, channel_table


def read_system_id_data(file):
    """Read system-identification data from a Rattlesnake npz file.

    Parses the FRF matrix, response CPSD, reference CPSD, response noise
    CPSD, reference noise CPSD, and multiple coherence from a numpy ``.npz``
    file produced by the Rattlesnake system-identification routine.  Applies
    response and reference transformation matrices when they are present in
    the file (i.e. when the corresponding array is not ``NaN``).

    Parameters
    ----------
    file : str or numpy.lib.npyio.NpzFile
        Path to the Rattlesnake ``.npz`` system-ID file, or an already-loaded
        ``NpzFile`` object.

    Returns
    -------
    frfs : TransferFunctionArray
        Frequency response function matrix from the system ID.
    response_cpsd : PowerSpectralDensityArray
        Response cross-power spectral density array.
    reference_cpsd : PowerSpectralDensityArray
        Reference cross-power spectral density array.
    response_noise_cpsd : PowerSpectralDensityArray
        Response noise cross-power spectral density array.
    reference_noise_cpsd : PowerSpectralDensityArray
        Reference noise cross-power spectral density array.
    coherence : MultipleCoherenceArray
        Multiple coherence array for each response channel.
    """

    if isinstance(file, str):
        file = np.load(file)
    df = file["sysid_frequency_spacing"]
    if np.isnan(file["response_transformation_matrix"]):
        try:
            response_dofs = coordinate_array(
                [int(v) for v in file["channel_node_number"][file["response_indices"]]],
                file["channel_node_direction"][file["response_indices"]],
            )
        except Exception:
            response_dofs = coordinate_array(file["response_indices"] + 1, 0)
    else:
        response_dofs = coordinate_array(
            np.arange(file["response_transformation_matrix"].shape[0]) + 1, 0
        )
    if np.isnan(file["reference_transformation_matrix"]):
        try:
            reference_dofs = coordinate_array(
                [int(v) for v in file["channel_node_number"][file["reference_indices"]]],
                file["channel_node_direction"][file["reference_indices"]],
            )
        except Exception:
            reference_dofs = coordinate_array(file["reference_indices"] + 1, 0)
    else:
        reference_dofs = coordinate_array(
            np.arange(file["reference_transformation_matrix"].shape[0]) + 1, 0
        )
    ordinate = np.moveaxis(file["frf_data"], 0, -1)
    frfs = data_array(
        FunctionTypes.FREQUENCY_RESPONSE_FUNCTION,
        df * np.arange(ordinate.shape[-1]),
        ordinate,
        outer_product(response_dofs, reference_dofs),
    )
    ordinate = np.moveaxis(file["response_cpsd"], 0, -1)
    response_cpsd = data_array(
        FunctionTypes.POWER_SPECTRAL_DENSITY,
        df * np.arange(ordinate.shape[-1]),
        ordinate,
        outer_product(response_dofs, response_dofs),
    )
    ordinate = np.moveaxis(file["reference_cpsd"], 0, -1)
    reference_cpsd = data_array(
        FunctionTypes.POWER_SPECTRAL_DENSITY,
        df * np.arange(ordinate.shape[-1]),
        ordinate,
        outer_product(reference_dofs, reference_dofs),
    )
    ordinate = np.moveaxis(file["response_noise_cpsd"], 0, -1)
    response_noise_cpsd = data_array(
        FunctionTypes.POWER_SPECTRAL_DENSITY,
        df * np.arange(ordinate.shape[-1]),
        ordinate,
        outer_product(response_dofs, response_dofs),
    )
    ordinate = np.moveaxis(file["reference_noise_cpsd"], 0, -1)
    reference_noise_cpsd = data_array(
        FunctionTypes.POWER_SPECTRAL_DENSITY,
        df * np.arange(ordinate.shape[-1]),
        ordinate,
        outer_product(reference_dofs, reference_dofs),
    )
    ordinate = np.moveaxis(file["coherence"], 0, -1)
    coherence = data_array(
        FunctionTypes.MULTIPLE_COHERENCE,
        df * np.arange(ordinate.shape[-1]),
        ordinate,
        outer_product(response_dofs),
    )
    return frfs, response_cpsd, reference_cpsd, response_noise_cpsd, reference_noise_cpsd, coherence


def read_system_id_nc4(file, coordinate_override_column=None):
    """Read system-identification spectral data from a Rattlesnake nc4 file.

    Parses FRF, response CPSD, drive CPSD, noise CPSD, and coherence arrays
    from the first non-channel environment group found in *file*.  Applies
    response and reference transformation matrices when present in the
    environment group.

    Parameters
    ----------
    file : str or netCDF4.Dataset
        Path to a Rattlesnake nc4 file, or an already-open ``netCDF4.Dataset``.
    coordinate_override_column : str, optional
        Name of a channel-table column whose string values are parsed as
        coordinates.  When ``None`` (default) coordinates are assembled from
        the ``node_number`` and ``node_direction`` channel-table columns.

    Returns
    -------
    frfs : TransferFunctionArray
        Frequency response functions with shape ``(n_responses, n_drives)``.
    response_cpsd : PowerSpectralDensityArray
        Response cross-power spectral density matrix.
    drive_cpsd : PowerSpectralDensityArray
        Drive (reference) cross-power spectral density matrix.
    response_noise_cpsd : PowerSpectralDensityArray
        Response noise cross-power spectral density matrix.
    drive_noise_cpsd : PowerSpectralDensityArray
        Drive noise cross-power spectral density matrix.
    coherence : MultipleCoherenceArray
        Multiple coherence for each response channel.
    """
    if isinstance(file, str):
        ds = nc4.Dataset(file, "r")
    elif isinstance(file, nc4.Dataset):
        ds = file

    environment = [group for group in ds.groups if not group == "channels"][0]

    # Get the channels in the group
    if coordinate_override_column is None:
        nodes = [
            int("".join(char for char in node if char in "0123456789"))
            for node in ds["channels"]["node_number"]
        ]
        directions = np.array(ds["channels"]["node_direction"][:], dtype="<U3")
        coordinates = coordinate_array(nodes, directions)
    else:
        coordinates = coordinate_array(string_array=ds["channels"][coordinate_override_column])
    drives = ds["channels"]["feedback_device"][:] != ""

    # Cull down to just those in the environment
    environment_index = np.where(ds["environment_names"][:] == environment)[0][0]
    environment_channels = ds["environment_active_channels"][:, environment_index].astype(bool)

    drives = drives[environment_channels]
    coordinates = coordinates[environment_channels]

    control_indices = ds[environment]["control_channel_indices"][:]

    if "response_transformation_matrix" in ds[environment].variables:
        control_coordinates = coordinate_array(
            np.arange(ds[environment]["response_transformation_matrix"].shape[0]) + 1, 0
        )
        response_transform_comment1 = np.array(
            [
                f"Unknown :: Transformed Response {i}"
                for i in np.arange(ds[environment]["response_transformation_matrix"].shape[0]) + 1
            ],
            dtype="<U80",
        )
        response_transform_comment2 = np.array(
            [
                f"Transformed Response {i} :: Transformed Response {i}"
                for i in np.arange(ds[environment]["response_transformation_matrix"].shape[0]) + 1
            ],
            dtype="<U80",
        )
        response_transform_comment3 = np.array(
            [
                f"Transformed Response {i} :: Transformed Response {i}"
                for i in np.arange(ds[environment]["response_transformation_matrix"].shape[0]) + 1
            ],
            dtype="<U80",
        )
        response_transform_comment4 = np.array(
            [
                f"Transformed Response {i}"
                for i in np.arange(ds[environment]["response_transformation_matrix"].shape[0]) + 1
            ],
            dtype="<U80",
        )
        response_transform_comment5 = np.array(
            [
                f"Transformed Response {i}"
                for i in np.arange(ds[environment]["response_transformation_matrix"].shape[0]) + 1
            ],
            dtype="<U80",
        )
        control_indices = np.arange(ds[environment]["response_transformation_matrix"].shape[0])
    else:
        control_coordinates = coordinates[control_indices]

    if "reference_transformation_matrix" in ds[environment].variables:
        drive_coordinates = coordinate_array(
            np.arange(ds[environment]["reference_transformation_matrix"].shape[0]) + 1, 0
        )
        drive_transform_comment1 = np.array(
            [
                f"Unknown :: Transformed Drive {i}"
                for i in np.arange(ds[environment]["reference_transformation_matrix"].shape[0]) + 1
            ],
            dtype="<U80",
        )
        drive_transform_comment2 = np.array(
            [
                f"Transformed Drive {i} :: Transformed Drive {i}"
                for i in np.arange(ds[environment]["reference_transformation_matrix"].shape[0]) + 1
            ],
            dtype="<U80",
        )
        drive_transform_comment3 = np.array(
            [
                f"Transformed Drive {i} :: Transformed Drive {i}"
                for i in np.arange(ds[environment]["reference_transformation_matrix"].shape[0]) + 1
            ],
            dtype="<U80",
        )
        drive_transform_comment4 = np.array(
            [
                f"Transformed Drive {i}"
                for i in np.arange(ds[environment]["reference_transformation_matrix"].shape[0]) + 1
            ],
            dtype="<U80",
        )
        drive_transform_comment5 = np.array(
            [
                f"Transformed Drive {i}"
                for i in np.arange(ds[environment]["reference_transformation_matrix"].shape[0]) + 1
            ],
            dtype="<U80",
        )
        drives = np.ones(ds[environment]["reference_transformation_matrix"].shape[0], dtype=bool)
    else:
        drive_coordinates = coordinates[drives]

    # Load the spectral data
    frequency_spacing = ds.sample_rate / ds[environment].sysid_frame_size
    fft_lines = ds[environment].dimensions["sysid_fft_lines"].size
    frequencies = np.arange(fft_lines) * frequency_spacing

    frf_array = np.moveaxis(
        np.array(ds[environment]["frf_data_real"][:] + 1j * ds[environment]["frf_data_imag"][:]),
        0,
        -1,
    )

    response_cpsd_array = np.moveaxis(
        np.array(
            ds[environment]["response_cpsd_real"][:] + 1j * ds[environment]["response_cpsd_imag"][:]
        ),
        0,
        -1,
    )

    drive_cpsd_array = np.moveaxis(
        np.array(
            ds[environment]["reference_cpsd_real"][:]
            + 1j * ds[environment]["reference_cpsd_imag"][:]
        ),
        0,
        -1,
    )

    response_noise_cpsd_array = np.moveaxis(
        np.array(
            ds[environment]["response_noise_cpsd_real"][:]
            + 1j * ds[environment]["response_noise_cpsd_imag"][:]
        ),
        0,
        -1,
    )

    drive_noise_cpsd_array = np.moveaxis(
        np.array(
            ds[environment]["reference_noise_cpsd_real"][:]
            + 1j * ds[environment]["reference_noise_cpsd_imag"][:]
        ),
        0,
        -1,
    )

    coherence_array = np.moveaxis(np.array(ds[environment]["frf_coherence"][:]), 0, -1)

    response_coordinates_cpsd = outer_product(control_coordinates, control_coordinates)
    drive_coordinates_cpsd = outer_product(drive_coordinates, drive_coordinates)
    frf_coordinates = outer_product(control_coordinates, drive_coordinates)
    coherence_coordinates = control_coordinates[:, np.newaxis]

    comment1 = np.char.add(
        np.char.add(np.array(ds["channels"]["channel_type"][:], dtype="<U80"), np.array(" :: ")),
        np.array(ds["channels"]["unit"][:], dtype="<U80"),
    )
    comment2 = np.char.add(
        np.char.add(np.array(ds["channels"]["physical_device"][:], dtype="<U80"), np.array(" :: ")),
        np.array(ds["channels"]["physical_channel"][:], dtype="<U80"),
    )
    comment3 = np.char.add(
        np.char.add(np.array(ds["channels"]["feedback_device"][:], dtype="<U80"), np.array(" :: ")),
        np.array(ds["channels"]["feedback_channel"][:], dtype="<U80"),
    )
    comment4 = np.array(ds["channels"]["comment"][:], dtype="<U80")
    comment5 = np.array(ds["channels"]["make"][:], dtype="<U80")
    for key in ("model", "serial_number", "triax_dof"):
        comment5 = np.char.add(comment5, np.array(" "))
        comment5 = np.char.add(comment5, np.array(ds["channels"][key][:], dtype="<U80"))

    full_comment1 = comment1[environment_channels]
    full_comment2 = comment2[environment_channels]
    full_comment3 = comment3[environment_channels]
    full_comment4 = comment4[environment_channels]
    full_comment5 = comment5[environment_channels]

    if "response_transformation_matrix" in ds[environment].variables:
        comment1 = response_transform_comment1
        comment2 = response_transform_comment2
        comment3 = response_transform_comment3
        comment4 = response_transform_comment4
        comment5 = response_transform_comment5
    else:
        comment1 = full_comment1
        comment2 = full_comment2
        comment3 = full_comment3
        comment4 = full_comment4
        comment5 = full_comment5
    comment1_response_cpsd = np.empty(
        (response_coordinates_cpsd.shape[0], response_coordinates_cpsd.shape[1]),
        dtype=comment1.dtype,
    )
    comment2_response_cpsd = np.empty(
        (response_coordinates_cpsd.shape[0], response_coordinates_cpsd.shape[1]),
        dtype=comment1.dtype,
    )
    comment3_response_cpsd = np.empty(
        (response_coordinates_cpsd.shape[0], response_coordinates_cpsd.shape[1]),
        dtype=comment1.dtype,
    )
    comment4_response_cpsd = np.empty(
        (response_coordinates_cpsd.shape[0], response_coordinates_cpsd.shape[1]),
        dtype=comment1.dtype,
    )
    comment5_response_cpsd = np.empty(
        (response_coordinates_cpsd.shape[0], response_coordinates_cpsd.shape[1]),
        dtype=comment1.dtype,
    )
    comment1_coherence = np.empty(response_coordinates_cpsd.shape[0], dtype=comment1.dtype)
    comment2_coherence = np.empty(response_coordinates_cpsd.shape[0], dtype=comment1.dtype)
    comment3_coherence = np.empty(response_coordinates_cpsd.shape[0], dtype=comment1.dtype)
    comment4_coherence = np.empty(response_coordinates_cpsd.shape[0], dtype=comment1.dtype)
    comment5_coherence = np.empty(response_coordinates_cpsd.shape[0], dtype=comment1.dtype)
    for i, idx in enumerate(control_indices):
        comment1_coherence[i] = comment1[idx]
        comment2_coherence[i] = comment2[idx]
        comment3_coherence[i] = comment3[idx]
        comment4_coherence[i] = comment4[idx]
        comment5_coherence[i] = comment5[idx]
        for j, jdx in enumerate(control_indices):
            comment1_response_cpsd[i, j] = comment1[idx] + " // " + comment1[jdx]
            comment2_response_cpsd[i, j] = comment2[idx] + " // " + comment2[jdx]
            comment3_response_cpsd[i, j] = comment3[idx] + " // " + comment3[jdx]
            comment4_response_cpsd[i, j] = comment4[idx] + " // " + comment4[jdx]
            comment5_response_cpsd[i, j] = comment5[idx] + " // " + comment5[jdx]

    if "reference_transformation_matrix" in ds[environment].variables:
        comment1 = drive_transform_comment1
        comment2 = drive_transform_comment2
        comment3 = drive_transform_comment3
        comment4 = drive_transform_comment4
        comment5 = drive_transform_comment5
    else:
        comment1 = full_comment1
        comment2 = full_comment2
        comment3 = full_comment3
        comment4 = full_comment4
        comment5 = full_comment5
    comment1_drive_cpsd = np.empty(
        (drive_coordinates_cpsd.shape[0], drive_coordinates_cpsd.shape[1]), dtype=comment1.dtype
    )
    comment2_drive_cpsd = np.empty(
        (drive_coordinates_cpsd.shape[0], drive_coordinates_cpsd.shape[1]), dtype=comment1.dtype
    )
    comment3_drive_cpsd = np.empty(
        (drive_coordinates_cpsd.shape[0], drive_coordinates_cpsd.shape[1]), dtype=comment1.dtype
    )
    comment4_drive_cpsd = np.empty(
        (drive_coordinates_cpsd.shape[0], drive_coordinates_cpsd.shape[1]), dtype=comment1.dtype
    )
    comment5_drive_cpsd = np.empty(
        (drive_coordinates_cpsd.shape[0], drive_coordinates_cpsd.shape[1]), dtype=comment1.dtype
    )
    drive_indices = np.where(drives)[0]
    for i, idx in enumerate(drive_indices):
        for j, jdx in enumerate(drive_indices):
            comment1_drive_cpsd[i, j] = comment1[idx] + " // " + comment1[jdx]
            comment2_drive_cpsd[i, j] = comment2[idx] + " // " + comment2[jdx]
            comment3_drive_cpsd[i, j] = comment3[idx] + " // " + comment3[jdx]
            comment4_drive_cpsd[i, j] = comment4[idx] + " // " + comment4[jdx]
            comment5_drive_cpsd[i, j] = comment5[idx] + " // " + comment5[jdx]

    if "response_transformation_matrix" in ds[environment].variables:
        rcomment1 = response_transform_comment1
        rcomment2 = response_transform_comment2
        rcomment3 = response_transform_comment3
        rcomment4 = response_transform_comment4
        rcomment5 = response_transform_comment5
    else:
        rcomment1 = full_comment1
        rcomment2 = full_comment2
        rcomment3 = full_comment3
        rcomment4 = full_comment4
        rcomment5 = full_comment5
    if "reference_transformation_matrix" in ds[environment].variables:
        dcomment1 = drive_transform_comment1
        dcomment2 = drive_transform_comment2
        dcomment3 = drive_transform_comment3
        dcomment4 = drive_transform_comment4
        dcomment5 = drive_transform_comment5
    else:
        dcomment1 = full_comment1
        dcomment2 = full_comment2
        dcomment3 = full_comment3
        dcomment4 = full_comment4
        dcomment5 = full_comment5

    comment1_frf = np.empty(
        (frf_coordinates.shape[0], frf_coordinates.shape[1]), dtype=comment1.dtype
    )
    comment2_frf = np.empty(
        (frf_coordinates.shape[0], frf_coordinates.shape[1]), dtype=comment1.dtype
    )
    comment3_frf = np.empty(
        (frf_coordinates.shape[0], frf_coordinates.shape[1]), dtype=comment1.dtype
    )
    comment4_frf = np.empty(
        (frf_coordinates.shape[0], frf_coordinates.shape[1]), dtype=comment1.dtype
    )
    comment5_frf = np.empty(
        (frf_coordinates.shape[0], frf_coordinates.shape[1]), dtype=comment1.dtype
    )
    for i, idx in enumerate(control_indices):
        for j, jdx in enumerate(drive_indices):
            comment1_frf[i, j] = rcomment1[idx] + " // " + dcomment1[jdx]
            comment2_frf[i, j] = rcomment2[idx] + " // " + dcomment2[jdx]
            comment3_frf[i, j] = rcomment3[idx] + " // " + dcomment3[jdx]
            comment4_frf[i, j] = rcomment4[idx] + " // " + dcomment4[jdx]
            comment5_frf[i, j] = rcomment5[idx] + " // " + dcomment5[jdx]

    # Save the data to SDynpy objects
    response_cpsd = data_array(
        FunctionTypes.POWER_SPECTRAL_DENSITY,
        frequencies,
        response_cpsd_array,
        response_coordinates_cpsd,
        comment1_response_cpsd,
        comment2_response_cpsd,
        comment3_response_cpsd,
        comment4_response_cpsd,
        comment5_response_cpsd,
    )
    response_noise_cpsd = data_array(
        FunctionTypes.POWER_SPECTRAL_DENSITY,
        frequencies,
        response_noise_cpsd_array,
        response_coordinates_cpsd,
        comment1_response_cpsd,
        comment2_response_cpsd,
        comment3_response_cpsd,
        comment4_response_cpsd,
        comment5_response_cpsd,
    )
    drive_cpsd = data_array(
        FunctionTypes.POWER_SPECTRAL_DENSITY,
        frequencies,
        drive_cpsd_array,
        drive_coordinates_cpsd,
        comment1_drive_cpsd,
        comment2_drive_cpsd,
        comment3_drive_cpsd,
        comment4_drive_cpsd,
        comment5_drive_cpsd,
    )
    drive_noise_cpsd = data_array(
        FunctionTypes.POWER_SPECTRAL_DENSITY,
        frequencies,
        drive_noise_cpsd_array,
        drive_coordinates_cpsd,
        comment1_drive_cpsd,
        comment2_drive_cpsd,
        comment3_drive_cpsd,
        comment4_drive_cpsd,
        comment5_drive_cpsd,
    )
    frfs = data_array(
        FunctionTypes.FREQUENCY_RESPONSE_FUNCTION,
        frequencies,
        frf_array,
        frf_coordinates,
        comment1_frf,
        comment2_frf,
        comment3_frf,
        comment4_frf,
        comment5_frf,
    )
    coherence = data_array(
        FunctionTypes.MULTIPLE_COHERENCE,
        frequencies,
        coherence_array,
        coherence_coordinates,
        comment1_coherence,
        comment2_coherence,
        comment3_coherence,
        comment4_coherence,
        comment5_coherence,
    )

    return frfs, response_cpsd, drive_cpsd, response_noise_cpsd, drive_noise_cpsd, coherence


def read_random_spectral_data(file, coordinate_override_column=None):
    """Read random-vibration spectral data from a Rattlesnake nc4 file.

    Reads the response CPSD, specification CPSD, and drive CPSD arrays that
    are written to disk while a Rattlesnake Random environment is running
    (i.e. during active control, not from a system-identification run).
    Applies response and reference transformation matrices when present.

    Parameters
    ----------
    file : str or netCDF4.Dataset
        Path to a Rattlesnake nc4 file, or an already-open ``netCDF4.Dataset``.
    coordinate_override_column : str, optional
        Name of a channel-table column whose string values are parsed as
        coordinates.  When ``None`` (default) coordinates are assembled from
        the ``node_number`` and ``node_direction`` channel-table columns.

    Returns
    -------
    response_cpsd : PowerSpectralDensityArray
        Measured response cross-power spectral density matrix.
    spec_cpsd : PowerSpectralDensityArray
        Specification (target) cross-power spectral density matrix.
    drive_cpsd : PowerSpectralDensityArray
        Drive (output) cross-power spectral density matrix.
    """

    if isinstance(file, str):
        ds = nc4.Dataset(file, "r")
    elif isinstance(file, nc4.Dataset):
        ds = file

    environment = [group for group in ds.groups if not group == "channels"][0]

    # Get the channels in the group
    if coordinate_override_column is None:
        nodes = [
            int("".join(char for char in node if char in "0123456789"))
            for node in ds["channels"]["node_number"]
        ]
        directions = np.array(ds["channels"]["node_direction"][:], dtype="<U3")
        coordinates = coordinate_array(nodes, directions)
    else:
        coordinates = coordinate_array(string_array=ds["channels"][coordinate_override_column])
    drives = ds["channels"]["feedback_device"][:] != ""

    # Cull down to just those in the environment
    environment_index = np.where(ds["environment_names"][:] == environment)[0][0]
    environment_channels = ds["environment_active_channels"][:, environment_index].astype(bool)

    drives = drives[environment_channels]
    coordinates = coordinates[environment_channels]

    control_indices = ds[environment]["control_channel_indices"][:]

    if "response_transformation_matrix" in ds[environment].variables:
        control_coordinates = coordinate_array(
            np.arange(ds[environment]["response_transformation_matrix"].shape[0]) + 1, 0
        )
        response_transform_comment1 = np.array(
            [
                f"Unknown :: Transformed Response {i}"
                for i in np.arange(ds[environment]["response_transformation_matrix"].shape[0]) + 1
            ],
            dtype="<U80",
        )
        response_transform_comment2 = np.array(
            [
                f"Transformed Response {i} :: Transformed Response {i}"
                for i in np.arange(ds[environment]["response_transformation_matrix"].shape[0]) + 1
            ],
            dtype="<U80",
        )
        response_transform_comment3 = np.array(
            [
                f"Transformed Response {i} :: Transformed Response {i}"
                for i in np.arange(ds[environment]["response_transformation_matrix"].shape[0]) + 1
            ],
            dtype="<U80",
        )
        response_transform_comment4 = np.array(
            [
                f"Transformed Response {i}"
                for i in np.arange(ds[environment]["response_transformation_matrix"].shape[0]) + 1
            ],
            dtype="<U80",
        )
        response_transform_comment5 = np.array(
            [
                f"Transformed Response {i}"
                for i in np.arange(ds[environment]["response_transformation_matrix"].shape[0]) + 1
            ],
            dtype="<U80",
        )
        control_indices = np.arange(ds[environment]["response_transformation_matrix"].shape[0])
    else:
        control_coordinates = coordinates[control_indices]

    if "reference_transformation_matrix" in ds[environment].variables:
        drive_coordinates = coordinate_array(
            np.arange(ds[environment]["reference_transformation_matrix"].shape[0]) + 1, 0
        )
        drive_transform_comment1 = np.array(
            [
                f"Unknown :: Transformed Drive {i}"
                for i in np.arange(ds[environment]["reference_transformation_matrix"].shape[0]) + 1
            ],
            dtype="<U80",
        )
        drive_transform_comment2 = np.array(
            [
                f"Transformed Drive {i} :: Transformed Drive {i}"
                for i in np.arange(ds[environment]["reference_transformation_matrix"].shape[0]) + 1
            ],
            dtype="<U80",
        )
        drive_transform_comment3 = np.array(
            [
                f"Transformed Drive {i} :: Transformed Drive {i}"
                for i in np.arange(ds[environment]["reference_transformation_matrix"].shape[0]) + 1
            ],
            dtype="<U80",
        )
        drive_transform_comment4 = np.array(
            [
                f"Transformed Drive {i}"
                for i in np.arange(ds[environment]["reference_transformation_matrix"].shape[0]) + 1
            ],
            dtype="<U80",
        )
        drive_transform_comment5 = np.array(
            [
                f"Transformed Drive {i}"
                for i in np.arange(ds[environment]["reference_transformation_matrix"].shape[0]) + 1
            ],
            dtype="<U80",
        )
        drives = np.ones(ds[environment]["reference_transformation_matrix"].shape[0], dtype=bool)
    else:
        drive_coordinates = coordinates[drives]

    # Load the spectral data
    frequencies = np.array(ds[environment]["specification_frequency_lines"][:])

    spec_cpsd = np.moveaxis(
        np.array(
            ds[environment]["specification_cpsd_matrix_real"][:]
            + 1j * ds[environment]["specification_cpsd_matrix_imag"][:]
        ),
        0,
        -1,
    )

    response_cpsd = np.moveaxis(
        np.array(
            ds[environment]["response_cpsd_real"][:] + 1j * ds[environment]["response_cpsd_imag"][:]
        ),
        0,
        -1,
    )

    drive_cpsd = np.moveaxis(
        np.array(
            ds[environment]["drive_cpsd_real"][:] + 1j * ds[environment]["drive_cpsd_imag"][:]
        ),
        0,
        -1,
    )

    response_coordinates_cpsd = outer_product(control_coordinates, control_coordinates)
    drive_coordinates_cpsd = outer_product(drive_coordinates, drive_coordinates)

    comment1 = np.char.add(
        np.char.add(np.array(ds["channels"]["channel_type"][:], dtype="<U80"), np.array(" :: ")),
        np.array(ds["channels"]["unit"][:], dtype="<U80"),
    )
    comment2 = np.char.add(
        np.char.add(np.array(ds["channels"]["physical_device"][:], dtype="<U80"), np.array(" :: ")),
        np.array(ds["channels"]["physical_channel"][:], dtype="<U80"),
    )
    comment3 = np.char.add(
        np.char.add(np.array(ds["channels"]["feedback_device"][:], dtype="<U80"), np.array(" :: ")),
        np.array(ds["channels"]["feedback_channel"][:], dtype="<U80"),
    )
    comment4 = np.array(ds["channels"]["comment"][:], dtype="<U80")
    comment5 = np.array(ds["channels"]["make"][:], dtype="<U80")
    for key in ("model", "serial_number", "triax_dof"):
        comment5 = np.char.add(comment5, np.array(" "))
        comment5 = np.char.add(comment5, np.array(ds["channels"][key][:], dtype="<U80"))

    full_comment1 = comment1[environment_channels]
    full_comment2 = comment2[environment_channels]
    full_comment3 = comment3[environment_channels]
    full_comment4 = comment4[environment_channels]
    full_comment5 = comment5[environment_channels]

    if "response_transformation_matrix" in ds[environment].variables:
        comment1 = response_transform_comment1
        comment2 = response_transform_comment2
        comment3 = response_transform_comment3
        comment4 = response_transform_comment4
        comment5 = response_transform_comment5
    else:
        comment1 = full_comment1
        comment2 = full_comment2
        comment3 = full_comment3
        comment4 = full_comment4
        comment5 = full_comment5
    comment1_response_cpsd = np.empty(
        (response_coordinates_cpsd.shape[0], response_coordinates_cpsd.shape[1]),
        dtype=comment1.dtype,
    )
    comment2_response_cpsd = np.empty(
        (response_coordinates_cpsd.shape[0], response_coordinates_cpsd.shape[1]),
        dtype=comment1.dtype,
    )
    comment3_response_cpsd = np.empty(
        (response_coordinates_cpsd.shape[0], response_coordinates_cpsd.shape[1]),
        dtype=comment1.dtype,
    )
    comment4_response_cpsd = np.empty(
        (response_coordinates_cpsd.shape[0], response_coordinates_cpsd.shape[1]),
        dtype=comment1.dtype,
    )
    comment5_response_cpsd = np.empty(
        (response_coordinates_cpsd.shape[0], response_coordinates_cpsd.shape[1]),
        dtype=comment1.dtype,
    )
    for i, idx in enumerate(control_indices):
        for j, jdx in enumerate(control_indices):
            comment1_response_cpsd[i, j] = comment1[idx] + " // " + comment1[jdx]
            comment2_response_cpsd[i, j] = comment2[idx] + " // " + comment2[jdx]
            comment3_response_cpsd[i, j] = comment3[idx] + " // " + comment3[jdx]
            comment4_response_cpsd[i, j] = comment4[idx] + " // " + comment4[jdx]
            comment5_response_cpsd[i, j] = comment5[idx] + " // " + comment5[jdx]

    if "reference_transformation_matrix" in ds[environment].variables:
        comment1 = drive_transform_comment1
        comment2 = drive_transform_comment2
        comment3 = drive_transform_comment3
        comment4 = drive_transform_comment4
        comment5 = drive_transform_comment5
    else:
        comment1 = full_comment1
        comment2 = full_comment2
        comment3 = full_comment3
        comment4 = full_comment4
        comment5 = full_comment5
    comment1_drive_cpsd = np.empty(
        (drive_coordinates_cpsd.shape[0], drive_coordinates_cpsd.shape[1]), dtype=comment1.dtype
    )
    comment2_drive_cpsd = np.empty(
        (drive_coordinates_cpsd.shape[0], drive_coordinates_cpsd.shape[1]), dtype=comment1.dtype
    )
    comment3_drive_cpsd = np.empty(
        (drive_coordinates_cpsd.shape[0], drive_coordinates_cpsd.shape[1]), dtype=comment1.dtype
    )
    comment4_drive_cpsd = np.empty(
        (drive_coordinates_cpsd.shape[0], drive_coordinates_cpsd.shape[1]), dtype=comment1.dtype
    )
    comment5_drive_cpsd = np.empty(
        (drive_coordinates_cpsd.shape[0], drive_coordinates_cpsd.shape[1]), dtype=comment1.dtype
    )
    drive_indices = np.where(drives)[0]
    for i, idx in enumerate(drive_indices):
        for j, jdx in enumerate(drive_indices):
            comment1_drive_cpsd[i, j] = comment1[idx] + " // " + comment1[jdx]
            comment2_drive_cpsd[i, j] = comment2[idx] + " // " + comment2[jdx]
            comment3_drive_cpsd[i, j] = comment3[idx] + " // " + comment3[jdx]
            comment4_drive_cpsd[i, j] = comment4[idx] + " // " + comment4[jdx]
            comment5_drive_cpsd[i, j] = comment5[idx] + " // " + comment5[jdx]

    # Save the data to SDynpy objects
    response_cpsd = data_array(
        FunctionTypes.POWER_SPECTRAL_DENSITY,
        frequencies,
        response_cpsd,
        response_coordinates_cpsd,
        comment1_response_cpsd,
        comment2_response_cpsd,
        comment3_response_cpsd,
        comment4_response_cpsd,
        comment5_response_cpsd,
    )
    spec_cpsd = data_array(
        FunctionTypes.POWER_SPECTRAL_DENSITY,
        frequencies,
        spec_cpsd,
        response_coordinates_cpsd,
        comment1_response_cpsd,
        comment2_response_cpsd,
        comment3_response_cpsd,
        comment4_response_cpsd,
        comment5_response_cpsd,
    )
    drive_cpsd = data_array(
        FunctionTypes.POWER_SPECTRAL_DENSITY,
        frequencies,
        drive_cpsd,
        drive_coordinates_cpsd,
        comment1_drive_cpsd,
        comment2_drive_cpsd,
        comment3_drive_cpsd,
        comment4_drive_cpsd,
        comment5_drive_cpsd,
    )
    return response_cpsd, spec_cpsd, drive_cpsd


def read_modal_data(file, coordinate_override_column=None, read_only_indices=None):
    """Read modal test data from a Rattlesnake nc4 file.

    Reads the full time history, FRF matrix, and multiple coherence array
    from a Rattlesnake modal-test nc4 file.  Time data are reshaped into
    ``(n_averages, n_channels, samples_per_frame)`` blocks.

    Parameters
    ----------
    file : str or netCDF4.Dataset
        Path to a Rattlesnake nc4 file, or an already-open ``netCDF4.Dataset``.
    coordinate_override_column : str, optional
        Channel-table column whose string values are parsed as coordinates.
        When ``None`` (default) coordinates are built from ``node_number`` and
        ``node_direction``.
    read_only_indices : slice or array_like, optional
        Index expression applied to the channel axis when loading time data.
        Defaults to ``slice(None)`` (all channels).

    Returns
    -------
    time_data : TimeHistoryArray
        Averaged time blocks with shape ``(n_averages, n_channels)``.
    frf_data : TransferFunctionArray
        Frequency response function matrix.
    coherence_data : MultipleCoherenceArray
        Multiple coherence for each response channel.
    channel_table : pandas.DataFrame
        Full channel table from the nc4 file.

    Warns
    -----
    UserWarning
        If the number of complete averages in the time data does not match the
        ``num_averages`` attribute stored in the test settings group.
    """
    if isinstance(file, str):
        ds = nc4.Dataset(file, "r")
    elif isinstance(file, nc4.Dataset):
        ds = file
    if read_only_indices is None:
        read_only_indices = slice(None)
    # Get parameters
    num_channels = ds.groups["channels"].variables["physical_device"].size
    group_key = [g for g in ds.groups if not g == "channels"][0]
    group = ds.groups[group_key]
    sample_rate = ds.sample_rate
    samples_per_frame = group.samples_per_frame
    num_averages = group.num_averages
    # Load in the time data
    try:
        output_data = (
            np.array(ds["time_data"][...][read_only_indices])
            .reshape(num_channels, num_averages, samples_per_frame)
            .transpose(1, 0, 2)
        )
    except ValueError:
        warnings.warn(
            "Number of averages in the time data does not match the number of averages specified in the test settings.  Your test may be incomplete."
        )
        output_data = (
            np.array(ds["time_data"][...][read_only_indices])
            .reshape(num_channels, -1, samples_per_frame)
            .transpose(1, 0, 2)
        )
    abscissa = np.arange(samples_per_frame) / sample_rate
    if coordinate_override_column is None:
        nodes = [
            int("".join(char for char in node if char in "0123456789"))
            for node in ds["channels"]["node_number"][...][read_only_indices]
        ]
        directions = np.array(ds["channels"]["node_direction"][...][read_only_indices], dtype="<U3")
        coordinates = coordinate_array(nodes, directions)[:, np.newaxis]
    else:
        coordinates = coordinate_array(
            string_array=ds["channels"][coordinate_override_column][read_only_indices]
        )[:, np.newaxis]
    array = {name: np.array(variable[:]) for name, variable in ds["channels"].variables.items()}
    channel_table = pd.DataFrame(array)
    comment1 = np.char.add(
        np.char.add(
            np.array(ds["channels"]["channel_type"][...][read_only_indices], dtype="<U80"),
            np.array(" :: "),
        ),
        np.array(ds["channels"]["unit"][...][read_only_indices], dtype="<U80"),
    )
    comment2 = np.char.add(
        np.char.add(
            np.array(ds["channels"]["physical_device"][...][read_only_indices], dtype="<U80"),
            np.array(" :: "),
        ),
        np.array(ds["channels"]["physical_channel"][...][read_only_indices], dtype="<U80"),
    )
    comment3 = np.char.add(
        np.char.add(
            np.array(ds["channels"]["feedback_device"][...][read_only_indices], dtype="<U80"),
            np.array(" :: "),
        ),
        np.array(ds["channels"]["feedback_channel"][...][read_only_indices], dtype="<U80"),
    )
    comment4 = np.array(ds["channels"]["comment"][...][read_only_indices], dtype="<U80")
    comment5 = np.array(ds["channels"]["make"][...][read_only_indices], dtype="<U80")
    for key in ("model", "serial_number", "triax_dof"):
        comment5 = np.char.add(comment5, np.array(" "))
        comment5 = np.char.add(
            comment5, np.array(ds["channels"][key][...][read_only_indices], dtype="<U80")
        )
    time_data = data_array(
        FunctionTypes.TIME_RESPONSE,
        abscissa,
        output_data,
        coordinates,
        comment1,
        comment2,
        comment3,
        comment4,
        comment5,
    )
    # Response and Reference Indices
    kept_indices = np.arange(num_channels)[read_only_indices]
    reference_indices = np.array(group.variables["reference_channel_indices"][:])
    response_indices = np.array(group.variables["response_channel_indices"][:])
    keep_response_indices = np.array(
        [i for i, index in enumerate(response_indices) if index in kept_indices]
    )
    keep_reference_indices = np.array(
        [i for i, index in enumerate(reference_indices) if index in kept_indices]
    )
    frequency_lines = (
        np.arange(group.dimensions["fft_lines"].size) * sample_rate / samples_per_frame
    )
    coherence_data = np.array(group["coherence"][:, keep_response_indices]).T
    comment1 = np.char.add(
        np.char.add(
            np.array(
                ds["channels"]["channel_type"][...][response_indices[keep_response_indices]],
                dtype="<U80",
            ),
            np.array(" :: "),
        ),
        np.array(
            ds["channels"]["unit"][...][response_indices[keep_response_indices]], dtype="<U80"
        ),
    )
    comment2 = np.char.add(
        np.char.add(
            np.array(
                ds["channels"]["physical_device"][...][response_indices[keep_response_indices]],
                dtype="<U80",
            ),
            np.array(" :: "),
        ),
        np.array(
            ds["channels"]["physical_channel"][...][response_indices[keep_response_indices]],
            dtype="<U80",
        ),
    )
    comment3 = np.char.add(
        np.char.add(
            np.array(
                ds["channels"]["feedback_device"][...][response_indices[keep_response_indices]],
                dtype="<U80",
            ),
            np.array(" :: "),
        ),
        np.array(
            ds["channels"]["feedback_channel"][...][response_indices[keep_response_indices]],
            dtype="<U80",
        ),
    )
    comment4 = np.array(
        ds["channels"]["comment"][...][response_indices[keep_response_indices]], dtype="<U80"
    )
    comment5 = np.array(
        ds["channels"]["make"][...][response_indices[keep_response_indices]], dtype="<U80"
    )
    for key in ("model", "serial_number", "triax_dof"):
        comment5 = np.char.add(comment5, np.array(" "))
        comment5 = np.char.add(
            comment5,
            np.array(
                ds["channels"][key][...][response_indices[keep_response_indices]], dtype="<U80"
            ),
        )
    coherence_data = data_array(
        FunctionTypes.MULTIPLE_COHERENCE,
        frequency_lines,
        coherence_data,
        coordinates[response_indices[keep_response_indices]],
        comment1,
        comment2,
        comment3,
        comment4,
        comment5,
    )
    # Frequency Response Functions
    frf_data = np.moveaxis(
        np.array(group["frf_data_real"])[
            :, keep_response_indices[:, np.newaxis], keep_reference_indices
        ]
        + np.array(group["frf_data_imag"])[
            :, keep_response_indices[:, np.newaxis], keep_reference_indices
        ]
        * 1j,
        0,
        -1,
    )
    frf_coordinate = outer_product(
        coordinates[response_indices[keep_response_indices], 0],
        coordinates[reference_indices[keep_reference_indices], 0],
    )
    # print(response_indices[keep_response_indices])
    # print(reference_indices[keep_reference_indices])
    response_comment1 = np.char.add(
        np.char.add(
            np.array(
                ds["channels"]["channel_type"][...][response_indices[keep_response_indices]],
                dtype="<U80",
            ),
            np.array(" :: "),
        ),
        np.array(
            ds["channels"]["unit"][...][response_indices[keep_response_indices]], dtype="<U80"
        ),
    )
    response_comment2 = np.char.add(
        np.char.add(
            np.array(
                ds["channels"]["physical_device"][...][response_indices[keep_response_indices]],
                dtype="<U80",
            ),
            np.array(" :: "),
        ),
        np.array(
            ds["channels"]["physical_channel"][...][response_indices[keep_response_indices]],
            dtype="<U80",
        ),
    )
    response_comment3 = np.char.add(
        np.char.add(
            np.array(
                ds["channels"]["feedback_device"][...][response_indices[keep_response_indices]],
                dtype="<U80",
            ),
            np.array(" :: "),
        ),
        np.array(
            ds["channels"]["feedback_channel"][...][response_indices[keep_response_indices]],
            dtype="<U80",
        ),
    )
    response_comment4 = np.array(
        ds["channels"]["comment"][...][response_indices[keep_response_indices]], dtype="<U80"
    )
    response_comment5 = np.array(
        ds["channels"]["make"][...][response_indices[keep_response_indices]], dtype="<U80"
    )
    for key in ("model", "serial_number", "triax_dof"):
        response_comment5 = np.char.add(response_comment5, np.array(" "))
        response_comment5 = np.char.add(
            response_comment5,
            np.array(
                ds["channels"][key][...][response_indices[keep_response_indices]], dtype="<U80"
            ),
        )
    reference_comment1 = np.char.add(
        np.char.add(
            np.array(
                ds["channels"]["channel_type"][...][reference_indices[keep_reference_indices]],
                dtype="<U80",
            ),
            np.array(" :: "),
        ),
        np.array(
            ds["channels"]["unit"][...][reference_indices[keep_reference_indices]], dtype="<U80"
        ),
    )
    reference_comment2 = np.char.add(
        np.char.add(
            np.array(
                ds["channels"]["physical_device"][...][reference_indices[keep_reference_indices]],
                dtype="<U80",
            ),
            np.array(" :: "),
        ),
        np.array(
            ds["channels"]["physical_channel"][...][reference_indices[keep_reference_indices]],
            dtype="<U80",
        ),
    )
    reference_comment3 = np.char.add(
        np.char.add(
            np.array(
                ds["channels"]["feedback_device"][...][reference_indices[keep_reference_indices]],
                dtype="<U80",
            ),
            np.array(" :: "),
        ),
        np.array(
            ds["channels"]["feedback_channel"][...][reference_indices[keep_reference_indices]],
            dtype="<U80",
        ),
    )
    reference_comment4 = np.array(
        ds["channels"]["comment"][...][reference_indices[keep_reference_indices]], dtype="<U80"
    )
    reference_comment5 = np.array(
        ds["channels"]["make"][...][reference_indices[keep_reference_indices]], dtype="<U80"
    )
    for key in ("model", "serial_number", "triax_dof"):
        reference_comment5 = np.char.add(reference_comment5, np.array(" "))
        reference_comment5 = np.char.add(
            reference_comment5,
            np.array(
                ds["channels"][key][...][reference_indices[keep_reference_indices]], dtype="<U80"
            ),
        )
    response_comment1, reference_comment1 = np.broadcast_arrays(
        response_comment1[:, np.newaxis], reference_comment1
    )
    comment1 = np.char.add(np.char.add(response_comment1, np.array(" / ")), reference_comment1)
    response_comment2, reference_comment2 = np.broadcast_arrays(
        response_comment2[:, np.newaxis], reference_comment2
    )
    comment2 = np.char.add(np.char.add(response_comment2, np.array(" / ")), reference_comment2)
    response_comment3, reference_comment3 = np.broadcast_arrays(
        response_comment3[:, np.newaxis], reference_comment3
    )
    comment3 = np.char.add(np.char.add(response_comment3, np.array(" / ")), reference_comment3)
    response_comment4, reference_comment4 = np.broadcast_arrays(
        response_comment4[:, np.newaxis], reference_comment4
    )
    comment4 = np.char.add(np.char.add(response_comment4, np.array(" / ")), reference_comment4)
    response_comment5, reference_comment5 = np.broadcast_arrays(
        response_comment5[:, np.newaxis], reference_comment5
    )
    comment5 = np.char.add(np.char.add(response_comment5, np.array(" / ")), reference_comment5)
    frf_data = data_array(
        FunctionTypes.FREQUENCY_RESPONSE_FUNCTION,
        frequency_lines,
        frf_data,
        frf_coordinate,
        comment1,
        comment2,
        comment3,
        comment4,
        comment5,
    )
    return time_data, frf_data, coherence_data, channel_table


def read_transient_control_data(file, coordinate_override_column=None):
    """Read transient control time-history data from a Rattlesnake nc4 file.

    Reads the control response signal, specification signal, and drive signal
    time histories from the first non-channel environment group in *file*.
    Applies response and reference transformation matrices when present.

    Parameters
    ----------
    file : str or netCDF4.Dataset
        Path to a Rattlesnake nc4 file, or an already-open ``netCDF4.Dataset``.
    coordinate_override_column : str, optional
        Channel-table column whose string values are parsed as coordinates.
        When ``None`` (default) coordinates are built from ``node_number`` and
        ``node_direction``.  Note: this argument is currently overridden to
        ``None`` inside the function body regardless of the value passed.

    Returns
    -------
    response_signal : TimeHistoryArray
        Measured control-response time history.
    spec_signal : TimeHistoryArray
        Specification (target) control signal time history.
    drive_signal : TimeHistoryArray
        Drive (output) signal time history.
    """
    if isinstance(file, str):
        ds = nc4.Dataset(file, "r")
    elif isinstance(file, nc4.Dataset):
        ds = file
    coordinate_override_column = None

    environment = [group for group in ds.groups if not group == "channels"][0]

    # Get the channels in the group
    if coordinate_override_column is None:
        nodes = [
            int("".join(char for char in node if char in "0123456789"))
            for node in ds["channels"]["node_number"]
        ]
        directions = np.array(ds["channels"]["node_direction"][:], dtype="<U3")
        coordinates = coordinate_array(nodes, directions)
    else:
        coordinates = coordinate_array(string_array=ds["channels"][coordinate_override_column])
    drives = ds["channels"]["feedback_device"][:] != ""

    # Cull down to just those in the environment
    environment_index = np.where(ds["environment_names"][:] == environment)[0][0]
    environment_channels = ds["environment_active_channels"][:, environment_index].astype(bool)

    drives = drives[environment_channels]
    coordinates = coordinates[environment_channels]

    control_indices = ds[environment]["control_channel_indices"][:]

    if "response_transformation_matrix" in ds[environment].variables:
        control_coordinates = coordinate_array(
            np.arange(ds[environment]["response_transformation_matrix"].shape[0]) + 1, 0
        )
        response_transform_comment1 = np.array(
            [
                f"Unknown :: Transformed Response {i}"
                for i in np.arange(ds[environment]["response_transformation_matrix"].shape[0]) + 1
            ],
            dtype="<U80",
        )
        response_transform_comment2 = np.array(
            [
                f"Transformed Response {i} :: Transformed Response {i}"
                for i in np.arange(ds[environment]["response_transformation_matrix"].shape[0]) + 1
            ],
            dtype="<U80",
        )
        response_transform_comment3 = np.array(
            [
                f"Transformed Response {i} :: Transformed Response {i}"
                for i in np.arange(ds[environment]["response_transformation_matrix"].shape[0]) + 1
            ],
            dtype="<U80",
        )
        response_transform_comment4 = np.array(
            [
                f"Transformed Response {i}"
                for i in np.arange(ds[environment]["response_transformation_matrix"].shape[0]) + 1
            ],
            dtype="<U80",
        )
        response_transform_comment5 = np.array(
            [
                f"Transformed Response {i}"
                for i in np.arange(ds[environment]["response_transformation_matrix"].shape[0]) + 1
            ],
            dtype="<U80",
        )
        control_indices = np.arange(ds[environment]["response_transformation_matrix"].shape[0])
    else:
        control_coordinates = coordinates[control_indices]

    if "reference_transformation_matrix" in ds[environment].variables:
        drive_coordinates = coordinate_array(
            np.arange(ds[environment]["reference_transformation_matrix"].shape[0]) + 1, 0
        )
        drive_transform_comment1 = np.array(
            [
                f"Unknown :: Transformed Drive {i}"
                for i in np.arange(ds[environment]["reference_transformation_matrix"].shape[0]) + 1
            ],
            dtype="<U80",
        )
        drive_transform_comment2 = np.array(
            [
                f"Transformed Drive {i} :: Transformed Drive {i}"
                for i in np.arange(ds[environment]["reference_transformation_matrix"].shape[0]) + 1
            ],
            dtype="<U80",
        )
        drive_transform_comment3 = np.array(
            [
                f"Transformed Drive {i} :: Transformed Drive {i}"
                for i in np.arange(ds[environment]["reference_transformation_matrix"].shape[0]) + 1
            ],
            dtype="<U80",
        )
        drive_transform_comment4 = np.array(
            [
                f"Transformed Drive {i}"
                for i in np.arange(ds[environment]["reference_transformation_matrix"].shape[0]) + 1
            ],
            dtype="<U80",
        )
        drive_transform_comment5 = np.array(
            [
                f"Transformed Drive {i}"
                for i in np.arange(ds[environment]["reference_transformation_matrix"].shape[0]) + 1
            ],
            dtype="<U80",
        )
        drives = np.ones(ds[environment]["reference_transformation_matrix"].shape[0], dtype=bool)
    else:
        drive_coordinates = coordinates[drives]

    # Load the time data
    timesteps = np.arange(ds[environment].dimensions["signal_samples"].size) / ds.sample_rate

    spec_signal = np.array(ds[environment]["control_signal"][...])

    response_signal = np.array(ds[environment]["control_response"][...])

    drive_signal = np.array(ds[environment]["control_drives"][...])

    response_coordinates = control_coordinates[:, np.newaxis]
    drive_coordinates = drive_coordinates[:, np.newaxis]

    comment1 = np.char.add(
        np.char.add(np.array(ds["channels"]["channel_type"][:], dtype="<U80"), np.array(" :: ")),
        np.array(ds["channels"]["unit"][:], dtype="<U80"),
    )
    comment2 = np.char.add(
        np.char.add(np.array(ds["channels"]["physical_device"][:], dtype="<U80"), np.array(" :: ")),
        np.array(ds["channels"]["physical_channel"][:], dtype="<U80"),
    )
    comment3 = np.char.add(
        np.char.add(np.array(ds["channels"]["feedback_device"][:], dtype="<U80"), np.array(" :: ")),
        np.array(ds["channels"]["feedback_channel"][:], dtype="<U80"),
    )
    comment4 = np.array(ds["channels"]["comment"][:], dtype="<U80")
    comment5 = np.array(ds["channels"]["make"][:], dtype="<U80")
    for key in ("model", "serial_number", "triax_dof"):
        comment5 = np.char.add(comment5, np.array(" "))
        comment5 = np.char.add(comment5, np.array(ds["channels"][key][:], dtype="<U80"))

    full_comment1 = comment1[environment_channels]
    full_comment2 = comment2[environment_channels]
    full_comment3 = comment3[environment_channels]
    full_comment4 = comment4[environment_channels]
    full_comment5 = comment5[environment_channels]

    if "response_transformation_matrix" in ds[environment].variables:
        comment1 = response_transform_comment1
        comment2 = response_transform_comment2
        comment3 = response_transform_comment3
        comment4 = response_transform_comment4
        comment5 = response_transform_comment5
    else:
        comment1 = full_comment1[control_indices]
        comment2 = full_comment2[control_indices]
        comment3 = full_comment3[control_indices]
        comment4 = full_comment4[control_indices]
        comment5 = full_comment5[control_indices]

    # Save the data to SDynpy objects
    response_signal = data_array(
        FunctionTypes.TIME_RESPONSE,
        timesteps,
        response_signal,
        response_coordinates,
        comment1,
        comment2,
        comment3,
        comment4,
        comment5,
    )
    spec_signal = data_array(
        FunctionTypes.TIME_RESPONSE,
        timesteps,
        spec_signal,
        response_coordinates,
        comment1,
        comment2,
        comment3,
        comment4,
        comment5,
    )

    if "reference_transformation_matrix" in ds[environment].variables:
        comment1 = drive_transform_comment1
        comment2 = drive_transform_comment2
        comment3 = drive_transform_comment3
        comment4 = drive_transform_comment4
        comment5 = drive_transform_comment5
    else:
        comment1 = full_comment1[drives]
        comment2 = full_comment2[drives]
        comment3 = full_comment3[drives]
        comment4 = full_comment4[drives]
        comment5 = full_comment5[drives]

    drive_signal = data_array(
        FunctionTypes.TIME_RESPONSE,
        timesteps,
        drive_signal,
        drive_coordinates,
        comment1,
        comment2,
        comment3,
        comment4,
        comment5,
    )

    return response_signal, spec_signal, drive_signal


def create_synthetic_test(
    spreadsheet_file_name: str,
    system_filename: str,
    system: System,
    excitation_coordinates: CoordinateArray,
    response_coordinates: CoordinateArray,
    rattlesnake_directory: str,
    displacement_derivative=2,
    sample_rate: int = None,
    time_per_read: float = None,
    time_per_write: float = None,
    integration_oversample: int = 10,
    environments: list = [],
    channel_comment_data: list = None,
    channel_serial_number_data: list = None,
    channel_triax_dof_data: list = None,
    channel_engineering_unit_data: list = None,
    channel_warning_level_data: list = None,
    channel_abort_level_data: list = None,
    channel_active_in_environment_data: dict = None,
):
    """Create a Rattlesnake synthetic-test spreadsheet and save the system model.

    Saves the structural dynamics *system* to *system_filename*, then uses the
    Rattlesnake ``components`` API to write a combined-environments profile
    template spreadsheet to *spreadsheet_file_name*.  The channel table and
    hardware sheets are populated with the supplied coordinates, sample rate,
    and optional per-channel data.

    Parameters
    ----------
    spreadsheet_file_name : str
        Path (including filename) for the output Rattlesnake Excel spreadsheet.
    system_filename : str
        Path where the ``System`` object is saved before populating the sheet.
    system : System
        Structural dynamics system model to save.
    excitation_coordinates : CoordinateArray
        Coordinates for excitation (force/drive) channels.
    response_coordinates : CoordinateArray
        Coordinates for response (acceleration/measurement) channels.
    rattlesnake_directory : str
        Directory containing the Rattlesnake ``components`` Python package.
    displacement_derivative : int, optional
        Order of derivative used when mapping from displacement to the measured
        quantity (e.g. 2 for acceleration).  Default is 2.
    sample_rate : int, optional
        Hardware sample rate in Hz written to the spreadsheet.  Default is
        ``None`` (left blank).
    time_per_read : float, optional
        Acquisition frame duration in seconds.  Default is ``None``.
    time_per_write : float, optional
        Output frame duration in seconds.  Default is ``None``.
    integration_oversample : int, optional
        Integration oversample factor written to the Hardware sheet.
        Default is 10.
    environments : list of (str, str), optional
        List of ``(environment_type, environment_name)`` tuples describing the
        environments to include in the profile template.  Default is ``[]``.
    channel_comment_data : list, optional
        Per-channel comment strings (column 4).  Default is ``None``.
    channel_serial_number_data : list, optional
        Per-channel serial number values (column 5).  Default is ``None``.
    channel_triax_dof_data : list, optional
        Per-channel triaxial DOF labels (column 6).  Default is ``None``.
    channel_engineering_unit_data : list, optional
        Per-channel engineering unit strings (column 8).  Default is ``None``.
    channel_warning_level_data : list, optional
        Per-channel warning level values (column 22).  Default is ``None``.
    channel_abort_level_data : list, optional
        Per-channel abort level values (column 23).  Default is ``None``.
    channel_active_in_environment_data : dict, optional
        Mapping of environment name to a boolean list indicating which channels
        are active in that environment.  When ``None`` (default) all channels
        are marked active in every environment.
    """
    system.save(system_filename)
    # Load in Rattlesnake to create a template for the test
    sys.path.insert(0, rattlesnake_directory)
    import components as rs

    environment_data = []
    for environment_type, environment_name in environments:
        # Find the identifier
        environment_type = rs.environments.ControlTypes[environment_type.upper()]
        environment_data.append((environment_type, environment_name))
    rs.ui_utilities.save_combined_environments_profile_template(
        spreadsheet_file_name, environment_data
    )
    sys.path.pop(0)
    # Populate the channel table
    workbook = opxl.load_workbook(spreadsheet_file_name)
    worksheet = workbook.get_sheet_by_name("Channel Table")
    index = 3
    for i, channel in enumerate(response_coordinates):
        worksheet.cell(index, 1, i + 1)
        worksheet.cell(index, 2, channel.node)
        worksheet.cell(index, 3, _string_map[channel.direction])
        worksheet.cell(index, 12, "Virtual")
        worksheet.cell(index, 14, "Accel")
        index += 1
    for i, channel in enumerate(excitation_coordinates):
        worksheet.cell(index, 1, len(response_coordinates) + i + 1)
        worksheet.cell(index, 2, channel.node)
        worksheet.cell(index, 3, _string_map[channel.direction])
        worksheet.cell(index, 12, "Virtual")
        worksheet.cell(index, 14, "Force")
        worksheet.cell(index, 20, "Shaker")
        index += 1
    # Go through the various channel table data that could have been optionally
    # provided
    for column, data in [
        (4, channel_comment_data),
        (5, channel_serial_number_data),
        (6, channel_triax_dof_data),
        (8, channel_engineering_unit_data),
        (22, channel_warning_level_data),
        (23, channel_abort_level_data),
    ]:
        if data is None:
            continue
        for row_index, value in enumerate(data):
            worksheet.cell(3 + row_index, column, value)
    # Now fill out the environment table
    if channel_active_in_environment_data is not None:
        for environment_index, (environment_type, environment_name) in enumerate(environment_data):
            for row_index, value in enumerate(channel_active_in_environment_data[environment_name]):
                if value:
                    worksheet.cell(3 + row_index, 24 + environment_index, "X")
    else:
        for environment_index, (environment_type, environment_name) in enumerate(environment_data):
            for row_index in range(response_coordinates.size + excitation_coordinates.size):
                worksheet.cell(3 + row_index, 24 + environment_index, "X")
    worksheet = workbook.get_sheet_by_name("Hardware")
    worksheet.cell(1, 2, 6)
    worksheet.cell(2, 2, os.path.abspath(system_filename))
    if sample_rate is not None:
        worksheet.cell(3, 2, sample_rate)
    if time_per_read is not None:
        worksheet.cell(4, 2, time_per_read)
    if time_per_write is not None:
        worksheet.cell(5, 2, time_per_write)
    worksheet.cell(6, 2, 1)
    worksheet.cell(7, 2, integration_oversample)
    workbook.save(spreadsheet_file_name)


def read_sine_control_data(
    control_file,
    read_quantities="control_response_signals_combined",
    excitation_dofs=None,
    control_dofs=None,
):
    """Read sine-control data from a Rattlesnake npz control file.

    Extracts one or more control quantities and returns them as
    ``TimeHistoryArray`` objects.  Quantities whose keys are suffixed with
    block indices (e.g. ``control_response_signals_combined_0``) are
    concatenated along the last axis before wrapping.

    Parameters
    ----------
    control_file : str or numpy.lib.npyio.NpzFile
        Path to a Rattlesnake ``.npz`` control file, or an already-loaded
        ``NpzFile`` object.
    read_quantities : str or list of str, optional
        Name or list of names of the quantity/quantities to extract.  Valid
        values are:

        *Concatenated* (built from numbered sub-keys):
          ``'control_response_signals_combined'``,
          ``'control_response_amplitudes'``,
          ``'control_response_phases'``,
          ``'control_drive_modifications'``

        *Unconcatenated* (single array):
          ``'control_response_frequencies'``,
          ``'control_response_arguments'``,
          ``'control_target_phases'``,
          ``'control_target_amplitudes'``

        Defaults to ``'control_response_signals_combined'``.
    excitation_dofs : CoordinateArray, optional
        Coordinate labels for excitation channels.  When ``None`` (default)
        sequential integer nodes with direction 0 are used.
    control_dofs : CoordinateArray, optional
        Coordinate labels for response/control channels.  When ``None``
        (default) sequential integer nodes with direction 0 are used.

    Returns
    -------
    TimeHistoryArray or list of TimeHistoryArray
        If *read_quantities* is a single string the return value is a single
        ``TimeHistoryArray``.  If it is a list, a list of
        ``TimeHistoryArray`` objects is returned in the same order.

    Raises
    ------
    ValueError
        If any entry in *read_quantities* is not one of the valid quantity
        names listed above.
    """
    concatenated_keys = [
        "control_response_signals_combined",
        "control_response_amplitudes",
        "control_response_phases",
        "control_drive_modifications",
    ]
    unconcatenated_keys = [
        "control_response_frequencies",
        "control_response_arguments",
        "control_target_phases",
        "control_target_amplitudes",
    ]
    dimension_labels = {}
    dimension_labels["control_response_signals_combined"] = ("response", "timestep")
    dimension_labels["control_response_amplitudes"] = ("tone", "response", "timestep")
    dimension_labels["control_response_phases"] = ("tone", "response", "timestep")
    dimension_labels["control_drive_modifications"] = ("tone", "excitation", "block_num")
    dimension_labels["achieved_excitation_signals_combined"] = ("excitation", "timestep")
    dimension_labels["achieved_excitation_signals"] = ("tone", "excitation", "timestep")
    dimension_labels["control_response_frequencies"] = ("tone", "timestep")
    dimension_labels["control_response_arguments"] = ("tone", "timestep")
    dimension_labels["control_target_amplitudes"] = ("tone", "response", "timestep")
    dimension_labels["control_target_phases"] = ("tone", "response", "timestep")
    if isinstance(control_file, str):
        control_file = np.load(control_file)
    sample_rate = control_file["sample_rate"]
    if isinstance(read_quantities, str):
        read_quantities = [read_quantities]
        return_single = True
    else:
        return_single = False
    return_data = []
    for read_quantity in read_quantities:
        try:
            dimension_label = dimension_labels[read_quantity]
        except KeyError:
            raise ValueError(
                f"{read_quantity} is not a valid quantity to read.  read_quantity must be one of {concatenated_keys+unconcatenated_keys}."
            )
        # Extract the data and concatenate if necessary
        if read_quantity in concatenated_keys:
            data = []
            for key in control_file:
                if read_quantity == "_".join(key.split("_")[:-1]):
                    this_data = control_file[key]
                    while this_data.ndim < len(dimension_label):
                        this_data = this_data[..., np.newaxis]
                    data.append(this_data)
            data = np.concatenate(data, axis=-1)
        elif read_quantity in unconcatenated_keys:
            data = control_file[read_quantity]
        else:
            raise ValueError(
                f"{read_quantity} is not a valid quantity to read.  read_quantity must be one of {concatenated_keys+unconcatenated_keys}."
            )
        # Set up the abscissa
        if dimension_label[-1] == "timestep":
            abscissa = np.arange(data.shape[-1]) / sample_rate
        elif dimension_label[-1] == "block_num":
            abscissa = np.arange(data.shape[-1])
        else:
            raise ValueError(f"{dimension_label[-1]} is an invalid entry.  How did you get here?")
        # Set up degrees of freedom
        if dimension_label[-2] == "response":
            if control_dofs is None:
                dofs = coordinate_array(np.arange(data.shape[-2]) + 1, 0)
            else:
                dofs = control_dofs
        elif dimension_label[-2] == "excitation":
            if excitation_dofs is None:
                dofs = coordinate_array(np.arange(data.shape[-2]) + 1, 0)
            else:
                dofs = excitation_dofs
        elif dimension_label[-2] == "tone":
            dofs = coordinate_array(np.arange(data.shape[-2]) + 1, 0)
        else:
            raise ValueError(f"{dimension_label[-2]} is an invalid entry.  How did you get here?")
        if any([dimension == "tone" for dimension in dimension_label]):
            comment1 = control_file["names"].reshape(
                *[-1 if dimension == "tone" else 1 for dimension in dimension_label][:-1]
            )
        else:
            comment1 = ""
        # Construct the TimeHistoryArray
        return_data.append(data_array(FunctionTypes.TIME_RESPONSE, abscissa, data, dofs, comment1))
    if return_single:
        return_data = return_data[0]
    return return_data


class RattlesnakeRandomEnvironmentData:
    """Data for a Rattlesnake Random vibration environment.

    Parses variables, dimensions, and attributes from the environment group of a
    Rattlesnake nc4 file and provides convenience properties that return
    sdynpy array objects.

    Parameters
    ----------
    channel_coordinates : CoordinateArray
        Coordinate array for all physical channels in the parent
        ``RattlesnakeData`` object.  Used to resolve control-channel
        coordinates without storing a back-reference to the parent.
    sysid_averages : int
        Number of averages used during system identification.
    samples_per_frame : int
        Number of time samples per data frame.
    cpsd_overlap : int
        Overlap count used when computing the CPSD.
    cpsd_window : str
        Window function name applied before CPSD computation.
    specification_frequency_lines : np.ndarray
        Frequency axis (Hz) for the specification CPSD matrices.
    specification_cpsd_matrix : np.ndarray
        Target CPSD specification matrix with shape
        ``(n_freq, n_control, n_control)``.
    specification_warning_matrix : np.ndarray
        Warning-band CPSD matrix with shape ``(2, n_freq, n_control, n_control)``
        where index 0 is the lower bound and index 1 is the upper bound.
    specification_abort_matrix : np.ndarray
        Abort-band CPSD matrix with shape ``(2, n_freq, n_control, n_control)``
        where index 0 is the lower bound and index 1 is the upper bound.
    control_channel_indices : np.ndarray
        Integer indices into the parent channel table identifying the control
        channels for this environment.
    fft_lines : int
        Number of FFT frequency lines used in control.
    **kwargs
        Any additional nc4 attributes, variables, or dimensions not covered
        by the explicit parameters above are stored directly as instance
        attributes via ``__setattr__``.
    """

    def __init__(
        self,
        channel_coordinates: CoordinateArray,
        sysid_averages: int,
        samples_per_frame: int,
        cpsd_overlap: int,
        cpsd_window: str,
        specification_frequency_lines: np.ndarray,
        specification_cpsd_matrix: np.ndarray,
        specification_warning_matrix: np.ndarray,
        specification_abort_matrix: np.ndarray,
        control_channel_indices: np.ndarray,
        fft_lines: int,
        sysid_frame_size: int = None,
        sysid_averaging_type: str = None,
        sysid_noise_averages: int = None,
        sysid_exponential_averaging_coefficient: float = None,
        sysid_estimator: str = None,
        sysid_level: float = None,
        sysid_level_ramp_time: float = None,
        sysid_signal_type: str = None,
        sysid_window: str = None,
        sysid_overlap: float = None,
        sysid_burst_on: float = None,
        sysid_pretrigger: float = None,
        sysid_burst_ramp_fraction: float = None,
        sysid_low_frequency_cutoff: float = None,
        sysid_high_frequency_cutoff: float = None,
        test_level_ramp_time: float = None,
        update_tf_during_control: int = None,
        cola_window: str = None,
        cola_overlap: float = None,
        cola_window_exponent: float = None,
        frames_in_cpsd: int = None,
        control_python_script: os.pathlike = None,
        control_python_function: str = None,
        control_python_function_type: int = None,
        control_python_function_parameters: str = None,
        allow_automatic_aborts: int = None,
        specification_channels: int = None,
        control_channels: int = None,
        **kwargs,
    ):
        """Store all environment parameters as instance attributes.

        Every named parameter is assigned directly to ``self``.  Any
        additional keyword arguments (extra nc4 attributes, variables, or
        dimensions) are stored via ``__setattr__``.
        """
        self.channel_coordinates = channel_coordinates
        self.sysid_averages = sysid_averages
        self.samples_per_frame = samples_per_frame
        self.cpsd_overlap = cpsd_overlap
        self.cpsd_window = cpsd_window.lower()
        self.specification_frequency_lines = specification_frequency_lines
        self.specification_cpsd_matrix = specification_cpsd_matrix
        self.specification_warning_matrix = specification_warning_matrix
        self.specification_abort_matrix = specification_abort_matrix
        self.control_channel_indices = control_channel_indices
        self.fft_lines = fft_lines
        self.sysid_frame_size = sysid_frame_size
        self.sysid_averaging_type = sysid_averaging_type
        self.sysid_noise_averages = sysid_noise_averages
        self.sysid_exponential_averaging_coefficient = sysid_exponential_averaging_coefficient
        self.sysid_estimator = sysid_estimator
        self.sysid_level = sysid_level
        self.sysid_level_ramp_time = sysid_level_ramp_time
        self.sysid_signal_type = sysid_signal_type
        self.sysid_window = sysid_window.lower()
        self.sysid_overlap = sysid_overlap
        self.sysid_burst_on = sysid_burst_on
        self.sysid_pretrigger = sysid_pretrigger
        self.sysid_burst_ramp_fraction = sysid_burst_ramp_fraction
        self.sysid_low_frequency_cutoff = sysid_low_frequency_cutoff
        self.sysid_high_frequency_cutoff = sysid_high_frequency_cutoff
        self.test_level_ramp_time = test_level_ramp_time
        self.update_tf_during_control = update_tf_during_control
        self.cola_window = cola_window
        self.cola_overlap = cola_overlap
        self.cola_window_exponent = cola_window_exponent
        self.frames_in_cpsd = frames_in_cpsd
        self.control_python_script = control_python_script
        self.control_python_function = control_python_function
        self.control_python_function_type = control_python_function_type
        self.control_python_function_parameters = control_python_function_parameters
        self.allow_automatic_aborts = allow_automatic_aborts
        self.specification_channels = specification_channels
        self.control_channels = control_channels
        for key, value in kwargs.items():
            self.__setattr__(key, value)

    @classmethod
    def load_from_env_group(
        cls, parent: "RattlesnakeData", env_group: nc4.Group
    ) -> "RattlesnakeRandomEnvironmentData":
        """Construct an instance by reading all data from an nc4 environment group.

        Loads nc4 attributes, variables (including complex split-variable pairs
        stored as ``<name>_real`` / ``<name>_imag``), and dimension sizes from
        *env_group*, then calls the constructor with
        ``channel_coordinates=parent.coordinate``.

        Parameters
        ----------
        parent : RattlesnakeData
            The top-level data object whose ``coordinate`` property provides
            the full physical-channel coordinate array.
        env_group : netCDF4.Group
            The nc4 group for this environment (e.g. ``data['Random']``).

        Returns
        -------
        RattlesnakeRandomEnvironmentData
            Fully populated environment data object.
        """

        kwargs = {}
        # Load environment attributes
        for attr in env_group.ncattrs():
            kwargs[attr] = env_group.__getattr__(attr)

        # Load environment variables (handle complex split-variable convention)
        loaded_vars = set()
        for var in list(env_group.variables.keys()):
            if var.endswith(("_real", "_imag")):
                base = var.removesuffix("_real").removesuffix("_imag")
                if base in loaded_vars:
                    continue
                else:
                    loaded_vars.add(base)
                    kwargs[base] = np.vectorize(complex)(
                        env_group[base + "_real"][:], env_group[base + "_imag"][:]
                    )
            else:
                loaded_vars.add(var)
                kwargs[var] = np.array(env_group[var])

        # Load environment dimensions
        for dim in env_group.dimensions:
            kwargs[dim] = env_group.dimensions[dim].size

        # Create the object
        obj = cls(channel_coordinates=parent.coordinate, **kwargs)

        return obj

    @property
    def control_coordinate(self) -> CoordinateArray:
        """Coordinate array for the control channels of this environment.

        When a ``response_transformation_matrix`` is present the coordinates
        are synthetic sequential nodes (1-based, direction 0) whose count
        equals the number of transformed virtual channels.  Otherwise
        ``channel_coordinates`` is indexed by ``control_channel_indices`` to
        return the physical control-channel coordinates.

        Returns
        -------
        CoordinateArray
            Coordinates for each control channel (or virtual transformed
            channel) in this environment.
        """
        if hasattr(self, "response_transformation_matrix"):
            control_coordinates = coordinate_array(
                node=np.arange(self.response_transformation_matrix.shape[0]) + 1, direction=0
            )
        else:
            control_coordinates = self.channel_coordinates[self.control_channel_indices]
        return control_coordinates

    @property
    def specification_cpsd(self) -> PowerSpectralDensityArray:
        """Target CPSD specification for this random environment.

        Assembles a square ``PowerSpectralDensityArray`` from the
        ``specification_cpsd_matrix`` and ``specification_frequency_lines``
        attributes loaded from the nc4 file.  Returns ``None`` when any of
        those attributes are absent.

        Returns
        -------
        PowerSpectralDensityArray or None
            Square cross-power spectral density array with shape
            ``(n_control, n_control)`` at each frequency line, or ``None``
            if the required attributes are not available.
        """
        if (
            hasattr(self, "specification_cpsd_matrix")
            and hasattr(self, "control_coordinate")
            and hasattr(self, "specification_frequency_lines")
        ):
            return power_spectral_density_array(
                abscissa=self.specification_frequency_lines,
                ordinate=np.moveaxis(self.specification_cpsd_matrix, 0, -1),
                coordinate=outer_product(self.control_coordinate, self.control_coordinate),
            )

    @property
    def specification_warning_psd(self) -> PowerSpectralDensityArray:
        """Upper and lower warning-band PSDs for this random environment.

        Assembles two ``PowerSpectralDensityArray`` objects from the lower
        (index 0) and upper (index -1) slices of ``specification_warning_matrix``
        along the first axis and concatenates them into a two-element array.
        Returns ``None`` when any required attribute is absent.

        Returns
        -------
        PowerSpectralDensityArray or None
            Array with shape ``(2, n_control)`` where index 0 is the lower
            warning bound and index 1 is the upper warning bound, or ``None``
            if ``specification_abort_matrix``, ``control_coordinate``, or
            ``specification_frequency_lines`` are not available.
        """
        if (
            hasattr(self, "specification_abort_matrix")
            and hasattr(self, "control_coordinate")
            and hasattr(self, "specification_frequency_lines")
        ):
            output_low = power_spectral_density_array(
                abscissa=self.specification_frequency_lines,
                ordinate=np.moveaxis(self.specification_warning_matrix[0, ...], 0, -1),
                coordinate=np.tile(self.control_coordinate, (2, 1)).T,
            )
            output_high = power_spectral_density_array(
                abscissa=self.specification_frequency_lines,
                ordinate=np.moveaxis(self.specification_warning_matrix[-1, ...], 0, -1),
                coordinate=np.tile(self.control_coordinate, (2, 1)).T,
            )
            return np.concatenate((output_low[np.newaxis, :], output_high[np.newaxis, :]))

    @property
    def specification_abort_psd(self) -> PowerSpectralDensityArray:
        """Upper and lower abort-band PSDs for this random environment.

        Assembles two ``PowerSpectralDensityArray`` objects from the lower
        (index 0) and upper (index -1) slices of ``specification_abort_matrix``
        along the first axis and concatenates them into a two-element array.
        Returns ``None`` when any required attribute is absent.

        Returns
        -------
        PowerSpectralDensityArray or None
            Array with shape ``(2, n_control)`` where index 0 is the lower
            abort bound and index 1 is the upper abort bound, or ``None`` if
            ``specification_warning_matrix``, ``control_coordinate``, or
            ``specification_frequency_lines`` are not available.
        """
        if (
            hasattr(self, "specification_warning_matrix")
            and hasattr(self, "control_coordinate")
            and hasattr(self, "specification_frequency_lines")
        ):
            output_low = power_spectral_density_array(
                abscissa=self.specification_frequency_lines,
                ordinate=np.moveaxis(self.specification_abort_matrix[0, ...], 0, -1),
                coordinate=np.tile(self.control_coordinate, (2, 1)).T,
            )
            output_high = power_spectral_density_array(
                abscissa=self.specification_frequency_lines,
                ordinate=np.moveaxis(self.specification_abort_matrix[-1, ...], 0, -1),
                coordinate=np.tile(self.control_coordinate, (2, 1)).T,
            )
            return np.concatenate((output_low[np.newaxis, :], output_high[np.newaxis, :]))


class RattlesnakeData:
    """Top-level container for all data read from a Rattlesnake nc4 file.

    Provides the global channel table, per-environment data objects, and
    convenience methods for assembling sdynpy array objects directly from the
    file data.

    Instances are normally created via :meth:`read_rattlesnake_nc4` rather
    than by calling the constructor directly.

    Parameters
    ----------
    sample_rate : int
        Acquisition sample rate in Hz.
    time_data : np.ndarray
        Raw time-domain data array loaded from the nc4 file.
    channel_table : pd.DataFrame
        Channel metadata table loaded from the ``channels`` group of the
        nc4 file.
    file_version : str, optional
        Version string stored in the nc4 file's global attributes.
    time_per_write : float, optional
        Duration of each write block in seconds.
    time_per_read : float, optional
        Duration of each read block in seconds.
    hardware : int, optional
        Hardware identifier stored in the nc4 file.
    hardware_file : str, optional
        Path to the hardware configuration file recorded in the nc4 file.
    output_oversample : int, optional
        Output oversample factor.
    task_trigger : int, optional
        Trigger channel index used to start/stop acquisition.
    task_trigger_output_channel : str, optional
        Name of the output channel used for task triggering.
    environment_names : list or np.ndarray, optional
        Ordered list of environment group names present in the nc4 file.
        Default is ``[]``.
    environment_active_channels : list or np.ndarray, optional
        Per-environment active channel masks.  Default is ``[]``.
    environments : dict, optional
        Mapping from environment name to the corresponding
        :class:`RattlesnakeRandomEnvironmentData` instance.  Populated by
        :meth:`read_rattlesnake_nc4`.  Default is ``{}``.
    **kwargs
        Any additional global nc4 attributes or variables not covered by
        the explicit parameters above are stored as instance attributes via
        ``__setattr__``.
    """

    # All known environment data classes, in order of preference.
    _ENV_CLASSES = (RattlesnakeRandomEnvironmentData,)

    def __init__(
        self,
        sample_rate: int,
        time_data: np.ndarray,
        channel_table: pd.DataFrame,
        file_version: str = None,
        time_per_write: float = None,
        time_per_read: float = None,
        hardware: int = None,
        hardware_file: str = None,
        output_oversample: int = None,
        task_trigger: int = None,
        task_trigger_output_channel: str = None,
        environment_names: list | np.ndarray = [],
        environment_active_channels: list | np.ndarray = [],
        environments: dict[str, RattlesnakeRandomEnvironmentData] = {},
        **kwargs,
    ):
        """Store all data parameters as instance attributes.

        Every named parameter is assigned directly to ``self``.  Any
        additional keyword arguments (extra global nc4 attributes or variables)
        are stored via ``__setattr__``.  See the class docstring for parameter
        descriptions.
        """
        self.sample_rate = sample_rate
        self.time_data = time_data
        self.channel_table = channel_table
        self.file_version = file_version
        self.time_per_write = time_per_write
        self.time_per_read = time_per_read
        self.hardware = hardware
        self.hardware_file = hardware_file
        self.output_oversample = output_oversample
        self.task_trigger = task_trigger
        self.task_trigger_output_channel = task_trigger_output_channel
        self.environment_names = environment_names
        self.environment_active_channels = environment_active_channels
        self.environments = environments
        for key, value in kwargs.items():
            self.__setattr__(key, value)

    @classmethod
    def read_rattlesnake_nc4(cls, filename: os.PathLike | nc4.Dataset):
        """Read a Rattlesnake nc4 file and return a populated data object.

        Loads global nc4 attributes, global variables (including all
        ``time_data*`` streams), the channel table, and each environment group
        into a new ``RattlesnakeData`` instance.  Each environment group is
        tried against every class in ``_ENV_CLASSES`` (currently only
        :class:`RattlesnakeRandomEnvironmentData`); the first class that loads
        without error wins.

        Parameters
        ----------
        filename : os.PathLike or netCDF4.Dataset
            Path to a Rattlesnake nc4 file, or an already-open
            ``netCDF4.Dataset``.

        Returns
        -------
        RattlesnakeData
            Fully populated data object with ``channel_table`` and
            ``environments`` attributes set.

        Raises
        ------
        Warning
            If environment groups are present in the file but none of the
            known environment classes can load any of them.
        """

        if isinstance(filename, nc4.Dataset):
            data = filename
        else:
            data = nc4.Dataset(filename)

        kwargs = {}
        # get global attributes
        for attr in data.ncattrs():
            kwargs[attr] = data.__getattr__(attr)

        # get global variables
        for var in list(data.variables.keys()):
            kwargs[var] = np.array(data.variables[var])

        # get channel table
        array = {
            name: np.array(variable[:])
            for name, variable in data.groups["channels"].variables.items()
        }
        kwargs["channel_table"] = pd.DataFrame(array)

        # create the object
        obj = cls(**kwargs)

        # get environment properties — detect which class to use for the environment - assign them to return object
        for env_name in obj.environment_names:
            env_group = data[env_name]
            for env_cls in cls._ENV_CLASSES:
                try:
                    obj.environments[env_name] = env_cls.load_from_env_group(
                        parent=obj, env_group=env_group
                    )
                    break  # first class that loads cleanly wins
                except Exception:
                    continue

        if not obj.environments:
            warnings.warn("None of the known environment classes could load any environment group from this file.")
        return obj

    @property
    def coordinate(self) -> CoordinateArray:
        """Coordinate array for all physical channels.

        Builds a ``CoordinateArray`` from the ``node_number`` and
        ``node_direction`` columns of the channel table.

        Returns
        -------
        CoordinateArray
            Coordinate array for all physical channels.
        """
        nodestrings = [
            "".join(char for char in node if char in "0123456789")
            for node in self.channel_table["node_number"]
        ]
        nodes = [i if node == "" else int(node) for i, node in enumerate(nodestrings)]
        directions = np.array(self.channel_table["node_direction"][:], dtype="<U3")
        return coordinate_array(nodes, directions)

    def get_coordinate(self, env_name: str = None) -> CoordinateArray:
        """Coordinate array for all channels, optionally including virtual channels.

        Parameters
        ----------
        env_name : str, optional
            Name of the environment to query for a response transformation
            matrix.  When ``None`` only the physical channel coordinates are
            returned.

        Returns
        -------
        CoordinateArray
            Physical channel coordinates, with virtual transformed-channel
            coordinates appended when *env_name* is supplied and a
            ``response_transformation_matrix`` is present.
        """
        coordinates = self.coordinate
        if env_name is not None and hasattr(
            self.environments[env_name], "response_transformation_matrix"
        ):
            virtual_coordinates = coordinate_array(
                node=np.arange(self.environments[env_name].response_transformation_matrix.shape[0])
                + 1,
                direction=0,
            )
            coordinates = np.concatenate([coordinates, virtual_coordinates])
        return coordinates

    @property
    def excitation_coordinate(self) -> CoordinateArray:
        """Coordinate array for excitation (drive) channels.

        Selects channels whose ``feedback_channel`` entry in the channel table
        is non-empty, which identifies them as drive/excitation channels.

        Returns
        -------
        CoordinateArray
            Subset of the full channel coordinate array containing only
            channels that have a non-empty ``feedback_channel`` value.
        """
        excitation_coordinate = self.coordinate[
            self.channel_table.feedback_channel.astype(str).str.strip().astype(bool)
        ]
        return excitation_coordinate

    @property
    def response_coordinate(self) -> CoordinateArray:
        """Coordinate array for response (non-excitation) channels.

        Returns all channel coordinates that are not present in
        :attr:`excitation_coordinate`.

        Returns
        -------
        CoordinateArray
            Subset of the full channel coordinate array containing only
            channels that do not have a non-empty ``feedback_channel`` value.
        """
        return np.setdiff1d(self.coordinate, self.excitation_coordinate)

    def units(self, env_name: str = None) -> np.ndarray:
        """Engineering-unit label for each channel, optionally including virtual channels.

        Reads the ``unit`` column of the channel table.  When *env_name* is
        provided and the named environment contains a
        ``response_transformation_matrix``, the array is zero-padded (empty
        string ``''``) for each virtual transformed channel.

        Parameters
        ----------
        env_name : str, optional
            Name of the environment to query for a response transformation
            matrix.  When ``None`` (default) only the physical channel units
            are returned.

        Returns
        -------
        np.ndarray of str
            Array of unit strings, one per channel (physical channels first,
            followed by empty-string placeholders for virtual transformed
            channels when *env_name* is supplied and a transformation matrix
            is present).
        """
        units = self.channel_table["unit"].astype(str).to_numpy()
        # Append Units for Virtual Channels of 'Transformed' responses
        if env_name is not None and hasattr(
            self.environments[env_name], "response_transformation_matrix"
        ):
            units = np.pad(
                units,
                (0, self.environments[env_name].response_transformation_matrix.shape[0]),
                constant_values="",
            ).astype(str)
        return units

    def get_time_data(
        self, index: int | None = None, env_name: str = None
    ) -> TimeHistoryArray | list[TimeHistoryArray]:
        """Build ``TimeHistoryArray`` objects from the source nc4 file data.

        Iterates over all ``time_data`` variables found on this data object
        (populated by :meth:`read_rattlesnake_nc4`), converts each to a
        ``TimeHistoryArray``, and optionally applies a response transformation
        matrix to produce virtual channel time histories which are appended to
        the array.

        Parameters
        ----------
        index : int or None, optional
            Zero-based index selecting a single time-data stream.  When
            ``None`` (default) all streams are returned as a list.
        env_name : str, optional
            Name of the environment to query for a ``response_transformation_matrix``.
            When provided and the environment has such a matrix, the
            transformed virtual-channel responses are computed and appended
            to each ``TimeHistoryArray``.  When ``None`` (default) no
            transformation is applied.

        Returns
        -------
        TimeHistoryArray or list of TimeHistoryArray or None
            A single ``TimeHistoryArray`` when *index* is specified, or a list
            of ``TimeHistoryArray`` objects (one per time-data stream) when
            *index* is ``None``.  Returns ``None`` if ``time_data`` or
            ``sample_rate`` attributes are not present on this object.
        """
        if hasattr(self, "time_data") and hasattr(self, "sample_rate"):
            names = [name for name in vars(self).keys() if name.startswith("time_data")]
            if index is not None:
                names = [names[index]]
            output = []
            for name in names:
                time_data_array = getattr(self, name)
                abscissa = np.arange(0, time_data_array.shape[-1]) / self.sample_rate
                time_data = time_history_array(
                    abscissa=abscissa,
                    ordinate=time_data_array,
                    coordinate=self.coordinate[:, np.newaxis],
                )

                if env_name is not None:
                    # If a Response Transformation is Used, Compute the Transformed Responses and append them to the end of the time_data with sequential nodes starting at 1 and increasing to the total number of transformed responses.
                    if hasattr(self.environments[env_name], "response_transformation_matrix"):
                        column_coordinates = self.coordinate[
                            self.environments[env_name].control_channel_indices
                        ]
                        row_coordinates = coordinate_array(
                            node=np.arange(
                                self.environments[env_name].response_transformation_matrix.shape[0]
                            )
                            + 1,
                            direction=0,
                        )
                        response_transformation = matrix(
                            matrix=self.environments[env_name].response_transformation_matrix,
                            row_coordinate=row_coordinates,
                            column_coordinate=column_coordinates,
                        )
                        time_data_transformed = time_data[
                            self.environments[env_name].control_channel_indices
                        ].apply_transformation(response_transformation)
                        time_data = np.concatenate([time_data, time_data_transformed])
                    output.append(time_data)
            return output if index is None else output[0]
